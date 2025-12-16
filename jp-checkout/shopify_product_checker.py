#!/usr/bin/env python3
"""
Shopify Page & Link Checker
A standalone tool to check for active pages and dead links in a Shopify store.

Usage:
    python shopify_product_checker.py <store_url>
    
Example:
    python shopify_product_checker.py https://mystore.myshopify.com
    python shopify_product_checker.py https://mystore.com
"""

import sys
import re
import json
import requests
from urllib.parse import urlparse, urljoin
from typing import List, Dict, Set, Optional
from dataclasses import dataclass
from html.parser import HTMLParser
import time

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  Warning: openpyxl not installed. Install with: pip install openpyxl")

@dataclass
class Page:
    """Page information"""
    id: str
    title: str
    handle: str
    url: str
    published: bool
    published_at: Optional[str]
    links: List[str]

@dataclass
class LinkCheck:
    """Link check result"""
    url: str
    status_code: Optional[int]
    is_dead: bool
    error: Optional[str]
    is_internal: bool

class LinkExtractor(HTMLParser):
    """Extract links from HTML content"""
    def __init__(self):
        super().__init__()
        self.links = []
    
    def handle_starttag(self, tag, attrs):
        if tag == 'a':
            for attr, value in attrs:
                if attr == 'href' and value:
                    self.links.append(value)

class ShopifyPageChecker:
    """Main checker class"""
    
    def __init__(self, store_url: str, access_token: Optional[str] = None):
        """
        Initialize the checker
        
        Args:
            store_url: Shopify store URL (e.g., https://mystore.myshopify.com or https://mystore.com)
            access_token: Optional Admin API access token for full page details
        """
        # Normalize store URL
        parsed = urlparse(store_url)
        if 'myshopify.com' in parsed.netloc:
            self.store_domain = parsed.netloc
            self.store_name = parsed.netloc.replace('.myshopify.com', '')
        else:
            # Try to extract store name from custom domain
            self.store_domain = parsed.netloc
            self.store_name = None
        
        self.base_url = f"{parsed.scheme}://{parsed.netloc}"
        self.storefront_api_url = f"{self.base_url}/api/2024-01/graphql.json"
        self.admin_api_url = f"https://{self.store_name}.myshopify.com/admin/api/2024-01" if self.store_name else None
        self.access_token = access_token
        
        self.pages: List[Page] = []
        self.all_links: Set[str] = set()
        self.link_checks: List[LinkCheck] = []
        
        # Domains to exclude (common external services)
        self.excluded_domains = [
            'google.com', 'googleapis.com', 'gstatic.com', 'googleusercontent.com',
            'shopify.com', 'shopifycdn.com', 'myshopify.com',
            'facebook.com', 'fb.com', 'facebook.net',
            'instagram.com',
            'twitter.com', 'x.com',
            'youtube.com', 'youtu.be',
            'pinterest.com',
            'linkedin.com',
            'tiktok.com',
            'snapchat.com',
            'whatsapp.com',
            'mailto:', 'tel:',
            'javascript:', '#',
        ]
    
    def get_pages_from_sitemap(self) -> List[Dict]:
        """Try to get pages from sitemap.xml"""
        pages = []
        sitemap_urls = [
            f"{self.base_url}/sitemap.xml",
            f"{self.base_url}/sitemap_pages.xml",
        ]
        
        for sitemap_url in sitemap_urls:
            try:
                print(f"🔍 Trying sitemap: {sitemap_url}")
                response = requests.get(sitemap_url, timeout=10)
                if response.status_code == 200:
                    # Parse XML sitemap
                    import xml.etree.ElementTree as ET
                    root = ET.fromstring(response.content)
                    
                    # Handle different sitemap formats
                    for url_elem in root.findall('.//{http://www.sitemaps.org/schemas/sitemap/0.9}url'):
                        loc_elem = url_elem.find('{http://www.sitemaps.org/schemas/sitemap/0.9}loc')
                        if loc_elem is not None:
                            url = loc_elem.text
                            # Check if it's a page URL
                            if '/pages/' in url:
                                handle = url.split('/pages/')[-1].rstrip('/')
                                if handle:
                                    pages.append({
                                        'id': f"page_{handle}",
                                        'title': handle.replace('-', ' ').title(),
                                        'handle': handle,
                                        'published': True,
                                        'publishedAt': None,
                                        'body': '',
                                        'bodySummary': ''
                                    })
                    
                    if pages:
                        print(f"✅ Found {len(pages)} pages in sitemap")
                        break
            except Exception as e:
                print(f"⚠️  Could not parse sitemap: {e}")
                continue
        
        return pages
    
    def get_pages_from_storefront_api(self) -> List[Dict]:
        """Fetch pages using Storefront API (public, no auth needed)"""
        pages = []
        cursor = None
        has_next_page = True
        
        query = """
        query getPages($cursor: String) {
          pages(first: 250, after: $cursor) {
            pageInfo {
              hasNextPage
              endCursor
            }
            edges {
              node {
                id
                title
                handle
                body
                bodySummary
                publishedAt
              }
            }
          }
        }
        """
        
        while has_next_page:
            variables = {"cursor": cursor} if cursor else {}
            
            try:
                response = requests.post(
                    self.storefront_api_url,
                    json={"query": query, "variables": variables},
                    headers={"Content-Type": "application/json"},
                    timeout=30
                )
                response.raise_for_status()
                data = response.json()
                
                if 'errors' in data:
                    print(f"⚠️  GraphQL Error: {data['errors']}")
                    break
                
                pages_data = data.get('data', {}).get('pages', {})
                edges = pages_data.get('edges', [])
                
                for edge in edges:
                    node = edge['node']
                    pages.append({
                        'id': node['id'],
                        'title': node['title'],
                        'handle': node['handle'],
                        'published': node.get('publishedAt') is not None,
                        'publishedAt': node.get('publishedAt'),
                        'body': node.get('body', ''),
                        'bodySummary': node.get('bodySummary', '')
                    })
                
                page_info = pages_data.get('pageInfo', {})
                has_next_page = page_info.get('hasNextPage', False)
                cursor = page_info.get('endCursor')
                
                print(f"📄 Fetched {len(pages)} pages so far...")
                time.sleep(0.5)  # Rate limiting
                
            except requests.exceptions.HTTPError as e:
                if e.response.status_code == 403:
                    print(f"⚠️  Storefront API access forbidden (403). Trying alternative methods...")
                    return []  # Return empty to try alternatives
                else:
                    print(f"❌ Error fetching pages: {e}")
                    return []
            except requests.exceptions.RequestException as e:
                print(f"⚠️  Error with Storefront API: {e}")
                return []
        
        return pages
    
    def scrape_page_content(self, page_handle: str) -> str:
        """Scrape page content directly from the store"""
        page_url = f"{self.base_url}/pages/{page_handle}"
        try:
            response = requests.get(page_url, timeout=10, headers={
                'User-Agent': 'Mozilla/5.0 (compatible; ShopifyChecker/1.0)'
            })
            if response.status_code == 200:
                return response.text
        except:
            pass
        return ""
    
    def get_storefront_pages(self) -> List[Dict]:
        """Try multiple methods to get pages"""
        print("🔍 Attempting to fetch pages...")
        
        # Method 1: Try Storefront API
        pages = self.get_pages_from_storefront_api()
        if pages:
            return pages
        
        # Method 2: Try sitemap
        print("🔍 Storefront API not available. Trying sitemap...")
        pages = self.get_pages_from_sitemap()
        if pages:
            return pages
        
        # Method 3: Try common page handles
        print("🔍 Trying common page handles...")
        common_handles = ['about', 'about-us', 'contact', 'privacy-policy', 'terms-of-service', 
                         'shipping', 'returns', 'faq', 'help', 'blog']
        
        found_pages = []
        for handle in common_handles:
            page_url = f"{self.base_url}/pages/{handle}"
            try:
                response = requests.head(page_url, timeout=5, allow_redirects=True)
                if response.status_code == 200:
                    # Try to get the page title
                    content = self.scrape_page_content(handle)
                    title = handle.replace('-', ' ').title()
                    if content:
                        # Try to extract title from HTML
                        title_match = re.search(r'<title[^>]*>([^<]+)</title>', content, re.IGNORECASE)
                        if title_match:
                            title = title_match.group(1).strip()
                    
                    found_pages.append({
                        'id': f"page_{handle}",
                        'title': title,
                        'handle': handle,
                        'published': True,
                        'publishedAt': None,
                        'body': content,
                        'bodySummary': ''
                    })
                    print(f"  ✅ Found: {title} ({handle})")
            except:
                continue
        
        if found_pages:
            print(f"✅ Found {len(found_pages)} pages using common handles")
            return found_pages
        
        return []
    
    def is_relevant_link(self, link: str) -> bool:
        """Check if a link is relevant (not a common external service)"""
        link_lower = link.lower()
        
        # Skip excluded domains
        for excluded in self.excluded_domains:
            if excluded in link_lower:
                return False
        
        # Skip mailto and tel links
        if link_lower.startswith('mailto:') or link_lower.startswith('tel:'):
            return False
        
        return True
    
    def extract_links_from_text(self, text: str) -> List[str]:
        """Extract links from HTML/text content, filtering out irrelevant ones"""
        links = []
        
        # Extract from HTML
        parser = LinkExtractor()
        parser.feed(text)
        links.extend(parser.links)
        
        # Also extract plain URLs
        url_pattern = r'https?://[^\s<>"{}|\\^`\[\]]+'
        plain_urls = re.findall(url_pattern, text)
        links.extend(plain_urls)
        
        # Clean and normalize links
        cleaned_links = []
        for link in links:
            link = link.strip()
            if link:
                # Convert relative URLs to absolute
                if link.startswith('/'):
                    link = urljoin(self.base_url, link)
                
                # Filter out irrelevant links
                if self.is_relevant_link(link):
                    cleaned_links.append(link)
        
        return list(set(cleaned_links))  # Remove duplicates
    
    def has_meaningful_content(self, html_content: str) -> bool:
        """Check if HTML has meaningful content (not just navigation/footer/cart)"""
        if not html_content:
            return False
        
        html_lower = html_content.lower()
        
        # Common indicators of empty/no-content pages
        empty_indicators = [
            'your cart is empty',
            'skip to content',
            'loading...',
            'estimated total',
            '₱0.00',
            '$0.00',
            'no content',
            'page not found',
            '404',
            'this page does not exist',
            'continue shopping',
            'check out',
            'log in to check out faster'
        ]
        
        # Check if page is mostly empty indicators (especially cart-related)
        empty_count = sum(1 for indicator in empty_indicators if indicator in html_lower)
        if empty_count >= 3:  # Multiple empty indicators suggest no content
            # But check if there's actual content beyond navigation
            pass
        
        # Extract text content (rough estimate)
        # Remove script and style tags
        text_content = re.sub(r'<script[^>]*>.*?</script>', '', html_content, flags=re.DOTALL | re.IGNORECASE)
        text_content = re.sub(r'<style[^>]*>.*?</style>', '', text_content, flags=re.DOTALL | re.IGNORECASE)
        # Remove common navigation/footer elements
        text_content = re.sub(r'<nav[^>]*>.*?</nav>', '', text_content, flags=re.DOTALL | re.IGNORECASE)
        text_content = re.sub(r'<footer[^>]*>.*?</footer>', '', text_content, flags=re.DOTALL | re.IGNORECASE)
        text_content = re.sub(r'<header[^>]*>.*?</header>', '', text_content, flags=re.DOTALL | re.IGNORECASE)
        
        # Extract text between tags
        text_only = re.sub(r'<[^>]+>', ' ', text_content)
        text_only = ' '.join(text_only.split())  # Normalize whitespace
        
        # Remove common navigation/footer text patterns
        navigation_patterns = [
            r'\bhome\b', r'\bshop\b', r'\bcart\b', r'\blogin\b', r'\bsearch\b', 
            r'\bmenu\b', r'\bfooter\b', r'\bheader\b', r'\bfacebook\b', 
            r'\binstagram\b', r'\btwitter\b', r'\byoutube\b', r'\btiktok\b',
            r'\babout\b', r'\bcontact\b', r'\bvisit our stores\b', r'\bstore hours\b',
            r'\bphone number\b', r'\bcustomer service\b', r'\bjoin our\b',
            r'\bpowered by shopify\b', r'\brefund policy\b', r'\bprivacy policy\b',
            r'\bterms of service\b', r'\bshipping policy\b', r'\b©\s*\d{4}',
            r'\bhave an account\b', r'\blog in\b', r'\bcontinue shopping\b',
            r'\bcheck out\b', r'\bestimated total\b', r'\btaxes.*calculated\b'
        ]
        
        meaningful_text = text_only
        for pattern in navigation_patterns:
            meaningful_text = re.sub(pattern, '', meaningful_text, flags=re.IGNORECASE)
        
        # Clean up extra whitespace
        meaningful_text = ' '.join(meaningful_text.split())
        
        # Look for meaningful content indicators
        content_indicators = [
            '<article',
            '<main',
            '<section',
            'class="content"',
            'class="post"',
            'class="article"',
            'class="blog-post"',
            'class="page-content"',
            'id="content"',
            'id="main-content"',
            'class="blog-content"',
            'class="post-content"',
        ]
        
        # Count content indicators
        content_count = sum(1 for indicator in content_indicators if indicator in html_lower)
        
        # Check for headings and paragraphs with actual text
        has_headings = bool(re.search(r'<h[1-6][^>]*>.*?</h[1-6]>', html_content, flags=re.DOTALL | re.IGNORECASE))
        has_paragraphs = bool(re.search(r'<p[^>]*>.*?</p>', html_content, flags=re.DOTALL | re.IGNORECASE))
        
        # Extract heading and paragraph text to verify they have content
        heading_text = ' '.join(re.findall(r'<h[1-6][^>]*>(.*?)</h[1-6]>', html_content, flags=re.DOTALL | re.IGNORECASE))
        paragraph_text = ' '.join(re.findall(r'<p[^>]*>(.*?)</p>', html_content, flags=re.DOTALL | re.IGNORECASE))
        
        # Clean extracted text
        heading_text = re.sub(r'<[^>]+>', ' ', heading_text)
        paragraph_text = re.sub(r'<[^>]+>', ' ', paragraph_text)
        heading_text = ' '.join(heading_text.split())
        paragraph_text = ' '.join(paragraph_text.split())
        
        # If we have content indicators and substantial text, it's meaningful
        if content_count >= 2 and len(meaningful_text.strip()) > 200:
            return True
        
        # If we have headings with actual text and paragraphs
        if has_headings and has_paragraphs:
            combined_content = heading_text + ' ' + paragraph_text
            # Remove navigation words from combined content
            for pattern in navigation_patterns:
                combined_content = re.sub(pattern, '', combined_content, flags=re.IGNORECASE)
            combined_content = ' '.join(combined_content.split())
            if len(combined_content.strip()) > 150:  # At least 150 chars of actual content
                return True
        
        # If we have substantial text even without specific content tags
        if len(meaningful_text.strip()) > 500:
            return True
        
        # Special case: Blog pages should have article/blog content
        is_blog_page = '/blog' in html_lower or 'blog' in html_lower
        if is_blog_page:
            # Blog pages need actual article content, not just navigation
            if len(meaningful_text.strip()) < 200:
                return False
            # Should have at least some headings or paragraphs with content
            if not has_headings and not has_paragraphs:
                return False
        
        return False
    
    def check_link(self, url: str, timeout: int = 10) -> LinkCheck:
        """Check if a link is dead or alive, including content validation"""
        is_internal = self.store_domain in url or url.startswith('/')
        
        try:
            # Use HEAD request first (faster)
            response = requests.head(
                url,
                timeout=timeout,
                allow_redirects=True,
                headers={'User-Agent': 'Mozilla/5.0 (compatible; ShopifyChecker/1.0)'}
            )
            status_code = response.status_code
            
            # If status is 200, check for meaningful content
            if status_code == 200:
                try:
                    # Fetch full page to check content
                    content_response = requests.get(
                        url,
                        timeout=timeout,
                        allow_redirects=True,
                        headers={'User-Agent': 'Mozilla/5.0 (compatible; ShopifyChecker/1.0)'}
                    )
                    
                    if content_response.status_code == 200:
                        # Check if page has meaningful content
                        if not self.has_meaningful_content(content_response.text):
                            return LinkCheck(
                                url, 
                                status_code, 
                                True, 
                                "No meaningful content (empty page)", 
                                is_internal
                            )
                except Exception as e:
                    # If content check fails, still consider it working (status 200)
                    pass
            
            is_dead = status_code >= 400
            
        except requests.exceptions.Timeout:
            return LinkCheck(url, None, True, "Timeout", is_internal)
        except requests.exceptions.ConnectionError:
            return LinkCheck(url, None, True, "Connection Error", is_internal)
        except requests.exceptions.RequestException as e:
            return LinkCheck(url, None, True, str(e), is_internal)
        except Exception as e:
            return LinkCheck(url, None, True, f"Unexpected error: {str(e)}", is_internal)
        
        return LinkCheck(url, status_code, is_dead, None, is_internal)
    
    def check_all_links(self, max_links: int = 100):
        """Check all extracted links"""
        links_to_check = list(self.all_links)[:max_links]
        total = len(links_to_check)
        
        print(f"\n🔍 Checking {total} links...")
        
        for i, link in enumerate(links_to_check, 1):
            print(f"  [{i}/{total}] Checking: {link[:60]}...", end='\r')
            check = self.check_link(link)
            self.link_checks.append(check)
            time.sleep(0.2)  # Rate limiting
        
        print()  # New line after progress
    
    def check_page_accessibility(self, page_handle: str) -> bool:
        """Check if a page is accessible (returns 200)"""
        page_url = f"{self.base_url}/pages/{page_handle}"
        try:
            response = requests.head(page_url, timeout=10, allow_redirects=True)
            return response.status_code == 200
        except:
            return False
    
    def analyze_pages(self):
        """Analyze all pages"""
        print("🔍 Analyzing pages...")
        
        pages_data = self.get_storefront_pages()
        
        if not pages_data:
            print("\n❌ No pages found or unable to fetch pages.")
            print("\n💡 Possible reasons:")
            print("  1. Storefront API is disabled or restricted")
            print("  2. No pages exist in the store")
            print("  3. Pages are password protected")
            print("\n💡 Solutions:")
            print("  - Enable Storefront API in Shopify Admin: Settings > Apps and sales channels > Develop apps")
            print("  - Or use Admin API with authentication (requires app setup)")
            return
        
        print(f"✅ Found {len(pages_data)} pages to analyze")
        
        for page_data in pages_data:
            # Extract links from page body
            body = page_data.get('body', '') or page_data.get('bodySummary', '')
            links = self.extract_links_from_text(body)
            self.all_links.update(links)
            
            # Check if page is published and accessible
            published = page_data.get('published', False)
            page_handle = page_data.get('handle', '')
            
            # Check if page is actually accessible
            is_accessible = False
            if published and page_handle:
                is_accessible = self.check_page_accessibility(page_handle)
            
            page = Page(
                id=page_data['id'],
                title=page_data['title'],
                handle=page_handle,
                url=f"{self.base_url}/pages/{page_handle}",
                published=published,
                published_at=page_data.get('publishedAt'),
                links=links
            )
            
            self.pages.append(page)
    
    def generate_report(self):
        """Generate a comprehensive report"""
        print("\n" + "="*80)
        print("📊 SHOPIFY PAGE & LINK CHECKER REPORT")
        print("="*80)
        
        # Page Summary
        total = len(self.pages)
        published = sum(1 for p in self.pages if p.published)
        unpublished = total - published
        
        print(f"\n📄 PAGE SUMMARY")
        print(f"  Total Pages: {total}")
        print(f"  ✅ Published (Live): {published}")
        print(f"  ❌ Unpublished/Draft: {unpublished}")
        
        # Link Summary
        total_links = len(self.all_links)
        checked_links = len(self.link_checks)
        dead_links = [lc for lc in self.link_checks if lc.is_dead]
        internal_links = [lc for lc in self.link_checks if lc.is_internal]
        external_links = [lc for lc in self.link_checks if not lc.is_internal]
        
        print(f"\n🔗 LINK SUMMARY")
        print(f"  Total Links Found: {total_links}")
        print(f"  Links Checked: {checked_links}")
        print(f"  ✅ Working Links: {checked_links - len(dead_links)}")
        print(f"  ❌ Dead Links: {len(dead_links)}")
        print(f"  🏠 Internal Links: {len(internal_links)}")
        print(f"  🌐 External Links: {len(external_links)}")
        
        # Dead Links Details
        if dead_links:
            print(f"\n❌ DEAD LINKS ({len(dead_links)})")
            for link_check in dead_links[:20]:  # Show first 20
                status_info = f"Status: {link_check.status_code}" if link_check.status_code else f"Error: {link_check.error}"
                print(f"  • {link_check.url}")
                print(f"    {status_info}")
            if len(dead_links) > 20:
                print(f"  ... and {len(dead_links) - 20} more")
        
        # Pages with Links
        pages_with_links = [p for p in self.pages if p.links]
        if pages_with_links:
            print(f"\n📎 PAGES WITH LINKS ({len(pages_with_links)})")
            for page in pages_with_links[:10]:  # Show first 10
                status_icon = "✅" if page.published else "❌"
                print(f"  {status_icon} {page.title}")
                print(f"    URL: {page.url}")
                print(f"    Status: {'Published' if page.published else 'Unpublished'}")
                print(f"    Links: {len(page.links)}")
                for link in page.links[:3]:  # Show first 3 links
                    print(f"      - {link}")
                if len(page.links) > 3:
                    print(f"      ... and {len(page.links) - 3} more")
            if len(pages_with_links) > 10:
                print(f"  ... and {len(pages_with_links) - 10} more pages")
        
        # Published Pages
        if published > 0:
            print(f"\n✅ PUBLISHED (LIVE) PAGES ({published})")
            for page in self.pages:
                if page.published:
                    print(f"  • {page.title}")
                    print(f"    URL: {page.url}")
                    if page.published_at:
                        print(f"    Published: {page.published_at}")
        
        # Unpublished Pages
        if unpublished > 0:
            print(f"\n❌ UNPUBLISHED/DRAFT PAGES ({unpublished})")
            for page in self.pages:
                if not page.published:
                    print(f"  • {page.title}")
                    print(f"    Handle: {page.handle}")
        
        # Save report to Excel file with website name
        domain_name = self.store_domain.replace('.myshopify.com', '').replace('.com', '').replace('.', '_')
        # Clean domain name for filename (remove invalid characters)
        domain_name = re.sub(r'[<>:"/\\|?*]', '_', domain_name)
        timestamp = int(time.time())
        report_file = f"shopify_pages_report_{domain_name}_{timestamp}.xlsx"
        self.generate_excel_report(report_file, total, published, unpublished, total_links, dead_links)
        
        print(f"\n💾 Excel report saved to: {report_file}")
        print("="*80)
    
    def generate_excel_report(self, filename: str, total: int, published: int, unpublished: int, 
                              total_links: int, dead_links: List[LinkCheck]):
        """Generate Excel report with formatted sheets"""
        if not OPENPYXL_AVAILABLE:
            # Fallback to JSON if openpyxl not available
            report_file = filename.replace('.xlsx', '.json')
            report_data = {
                'store_url': self.base_url,
                'timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
                'summary': {
                    'total_pages': total,
                    'published_pages': published,
                    'unpublished_pages': unpublished,
                    'total_links': total_links,
                    'dead_links': len(dead_links),
                },
                'pages': [
                    {
                        'title': p.title,
                        'handle': p.handle,
                        'url': p.url,
                        'published': p.published,
                        'published_at': p.published_at,
                        'links': p.links
                    }
                    for p in self.pages
                ],
                'dead_links': [
                    {
                        'url': lc.url,
                        'status_code': lc.status_code,
                        'error': lc.error,
                        'is_internal': lc.is_internal
                    }
                    for lc in dead_links
                ]
            }
            with open(report_file, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, ensure_ascii=False)
            print(f"⚠️  openpyxl not available. Saved JSON report instead: {report_file}")
            return
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Summary", 0)
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        
        ws_summary['A1'] = 'Shopify Page & Link Checker Report'
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:B1')
        
        ws_summary['A3'] = 'Store URL'
        ws_summary['B3'] = self.base_url
        ws_summary['A4'] = 'Report Date'
        ws_summary['B4'] = time.strftime('%Y-%m-%d %H:%M:%S')
        
        row = 6
        ws_summary[f'A{row}'] = 'Metric'
        ws_summary[f'B{row}'] = 'Value'
        ws_summary[f'A{row}'].fill = header_fill
        ws_summary[f'A{row}'].font = header_font
        ws_summary[f'B{row}'].fill = header_fill
        ws_summary[f'B{row}'].font = header_font
        ws_summary[f'A{row}'].border = border
        ws_summary[f'B{row}'].border = border
        
        summary_data = [
            ('Total Pages', total),
            ('Published (Live)', published),
            ('Unpublished/Draft', unpublished),
            ('Total Links Found', total_links),
            ('Dead Links', len(dead_links)),
            ('Working Links', total_links - len(dead_links)),
        ]
        
        for i, (label, value) in enumerate(summary_data, start=row+1):
            ws_summary[f'A{i}'] = label
            ws_summary[f'B{i}'] = value
            ws_summary[f'A{i}'].border = border
            ws_summary[f'B{i}'].border = border
        
        # Helper function to create a sheet with headers
        def create_sheet_with_headers(sheet_name, headers_list, col_widths):
            ws = wb.create_sheet(sheet_name)
            for col, header in enumerate(headers_list, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            for col, width in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(col)].width = width
            ws.freeze_panes = 'A2'
            return ws
        
        # Sheet 2: Published (Live) Pages
        active_pages = [p for p in self.pages if p.published]
        if active_pages:
            ws_active = create_sheet_with_headers(
                f"Published (Live) - {published}",
                ['Page Title', 'Page URL', 'Published Date', 'Links Count'],
                [35, 60, 20, 15]
            )
            
            for row_idx, page in enumerate(active_pages, start=2):
                published_date = page.published_at[:10] if page.published_at else "N/A"
                
                ws_active.cell(row=row_idx, column=1, value=page.title).border = border
                ws_active.cell(row=row_idx, column=2, value=page.url).border = border
                ws_active.cell(row=row_idx, column=3, value=published_date).border = border
                ws_active.cell(row=row_idx, column=4, value=len(page.links)).border = border
                
                # Make URL clickable
                ws_active.cell(row=row_idx, column=2).hyperlink = page.url
                ws_active.cell(row=row_idx, column=2).font = Font(color="0563C1", underline="single")
        
        # Sheet 3: Unpublished/Draft Pages
        unpublished_pages = [p for p in self.pages if not p.published]
        if unpublished_pages:
            ws_unpublished = create_sheet_with_headers(
                f"Unpublished/Draft - {unpublished}",
                ['Page Title', 'Handle', 'Links Count'],
                [35, 30, 15]
            )
            
            for row_idx, page in enumerate(unpublished_pages, start=2):
                ws_unpublished.cell(row=row_idx, column=1, value=page.title).border = border
                ws_unpublished.cell(row=row_idx, column=2, value=page.handle).border = border
                ws_unpublished.cell(row=row_idx, column=3, value=len(page.links)).border = border
        
        # Sheet 4: Total Links Found
        all_extracted_links = []
        for page in self.pages:
            for link in page.links:
                all_extracted_links.append({
                    'url': link,
                    'page_title': page.title,
                    'page_url': page.url
                })
        
        if all_extracted_links:
            ws_links = create_sheet_with_headers(
                f"Total Links Found - {total_links}",
                ['Link URL', 'Found In Page', 'Page URL', 'Type'],
                [60, 30, 50, 15]
            )
            
            for row_idx, link_data in enumerate(all_extracted_links, start=2):
                link_url = link_data['url']
                is_internal = self.store_domain in link_url or link_url.startswith(self.base_url)
                link_type = "Internal" if is_internal else "External"
                
                ws_links.cell(row=row_idx, column=1, value=link_url).border = border
                ws_links.cell(row=row_idx, column=2, value=link_data['page_title']).border = border
                ws_links.cell(row=row_idx, column=3, value=link_data['page_url']).border = border
                ws_links.cell(row=row_idx, column=4, value=link_type).border = border
                
                # Make URLs clickable
                ws_links.cell(row=row_idx, column=1).hyperlink = link_url
                ws_links.cell(row=row_idx, column=1).font = Font(color="0563C1", underline="single")
                ws_links.cell(row=row_idx, column=3).hyperlink = link_data['page_url']
                ws_links.cell(row=row_idx, column=3).font = Font(color="0563C1", underline="single")
        
        # Sheet 5: Dead Links
        if dead_links:
            ws_dead = create_sheet_with_headers(
                f"Dead Links - {len(dead_links)}",
                ['URL', 'Status Code', 'Error', 'Type'],
                [60, 15, 30, 15]
            )
            
            for row_idx, link_check in enumerate(dead_links, start=2):
                link_type = "Internal" if link_check.is_internal else "External"
                status_code = str(link_check.status_code) if link_check.status_code else "N/A"
                error = link_check.error or "N/A"
                
                ws_dead.cell(row=row_idx, column=1, value=link_check.url).border = border
                ws_dead.cell(row=row_idx, column=2, value=status_code).border = border
                ws_dead.cell(row=row_idx, column=3, value=error).border = border
                ws_dead.cell(row=row_idx, column=4, value=link_type).border = border
                
                # Make URL clickable
                ws_dead.cell(row=row_idx, column=1).hyperlink = link_check.url
                ws_dead.cell(row=row_idx, column=1).font = Font(color="0563C1", underline="single")
        
        # Sheet 6: Working Links
        working_links = [lc for lc in self.link_checks if not lc.is_dead]
        if working_links:
            ws_working = create_sheet_with_headers(
                f"Working Links - {len(working_links)}",
                ['URL', 'Status Code', 'Type'],
                [60, 15, 15]
            )
            
            for row_idx, link_check in enumerate(working_links, start=2):
                link_type = "Internal" if link_check.is_internal else "External"
                status_code = str(link_check.status_code) if link_check.status_code else "N/A"
                
                ws_working.cell(row=row_idx, column=1, value=link_check.url).border = border
                ws_working.cell(row=row_idx, column=2, value=status_code).border = border
                ws_working.cell(row=row_idx, column=3, value=link_type).border = border
                
                # Make URL clickable
                ws_working.cell(row=row_idx, column=1).hyperlink = link_check.url
                ws_working.cell(row=row_idx, column=1).font = Font(color="0563C1", underline="single")
        
        wb.save(filename)

def main():
    """Main entry point"""
    if len(sys.argv) < 2:
        print(__doc__)
        print("\n❌ Error: Store URL required")
        print("\nUsage:")
        print("  python shopify_product_checker.py <store_url>")
        print("\nExample:")
        print("  python shopify_product_checker.py https://mystore.myshopify.com")
        print("  python shopify_product_checker.py https://mystore.com")
        sys.exit(1)
    
    store_url = sys.argv[1]
    access_token = sys.argv[2] if len(sys.argv) > 2 else None
    
    print(f"🚀 Starting Shopify Page & Link Checker")
    print(f"📍 Store: {store_url}")
    print()
    
    checker = ShopifyPageChecker(store_url, access_token)
    
    try:
        # Analyze pages
        checker.analyze_pages()
        
        # Check links (optional - can be slow)
        if checker.all_links:
            response = input(f"\n🔍 Found {len(checker.all_links)} links. Check them for dead links? (y/n): ")
            if response.lower() == 'y':
                checker.check_all_links()
        
        # Generate report
        checker.generate_report()
        
    except KeyboardInterrupt:
        print("\n\n⚠️  Interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()

